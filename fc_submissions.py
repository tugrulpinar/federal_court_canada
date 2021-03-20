from selenium import webdriver
from selenium.webdriver.support.ui import Select
from data_base import Contacts, SubmissionTypes
from sub_decision_maker import SubDMaker
from bs4 import BeautifulSoup
from datetime import datetime
from pathlib import Path
import openpyxl
import shutil
import time
import os

source_dir = os.getcwd()
engine = SubDMaker()
engine.initiate()


# OPEN THE BROWSER
browser = webdriver.Chrome()
browser.get("https://efiling.fct-cf.gc.ca/en/online-access/e-filing-intro")
browser.maximize_window()
browser.implicitly_wait(20)


def find_click(element_id):
    variable = browser.find_element_by_id(element_id)
    variable.click()


def fill_out(element_id, data):
    variable = browser.find_element_by_id(element_id)
    variable.send_keys(data)


def drop_down_selection(element_id, selection):
    variable = browser.find_element_by_id(element_id)
    var = Select(variable)
    var.select_by_visible_text(selection)


def upload(element_id, document):
    variable = browser.find_element_by_id(element_id)
    doc_path = f"{document}.pdf"
    abs_path = os.path.join(source_dir, doc_path)
    variable.send_keys(abs_path)


def existing_filing():
    # AGREE TO THE TERMS
    find_click("disclamer")

    # FILE DOCUMENTS ON EXISTING PROCEEDING
    find_click("ExistingFiling")


def court_case(court_no):
    # STEP 1 - COURT NUMBER
    fill_out("userEnteredCourtNo", court_no)

    # NOT MUCH TO DO HERE. WE'RE JUS GOING TO CONTINUE
    for _ in range(2):
        find_click("Submit-Step-1")
        time.sleep(0.5)


def edit_party():
    # THERE'S NO NEED TO EDIT. WE'LL JUST CONTINUE
    find_click("Submit-Step-2")
    time.sleep(0.5)


def add_document():
    # CLICK ADD ROW BUTTON 3 TIMES. SELECT THE LANGUAGE AS ENGLISH.
    for x in range(3):
        find_click("addrow")
        document_language = browser.find_element_by_id(f"DocumentLanguage_{x}")
        document_language.click()
        english = Select(document_language)
        english.select_by_visible_text("English")

    # CHECK THE BOXES. WE WILL ONLY CHECK TWO OF THEM.
    for x in range(1, 3):
        checkbox = browser.find_element_by_id(f"filer_0_{x}")
        checkbox.click()
        checkbox2 = browser.find_element_by_id(f"filer_1_{x}")
        checkbox2.click()
        checkbox3 = browser.find_element_by_id(f"filer_2_{x}")
        checkbox3.click()


def upload_submission(type_of_submission):
    # CLICK THE DOCUMENT TYPE DROP DOWN LIST
    doc_type0 = browser.find_element_by_id("DocumentType_0")
    doc_type0.click()
    type_of_submission = type_of_submission.lower()

    # DEPENDING ON THE USER INPUT, SELECT THE SUBMISSION TYPE
    if type_of_submission == "a":
        app_record = Select(doc_type0)
        app_record.select_by_visible_text("IMM - APPLICATION RECORD")
        upload("file_0", "Application Record")
    elif type_of_submission == "b":
        reply_memo = Select(doc_type0)
        reply_memo.select_by_visible_text("IMM - REPLY MEMORANDUM")
        upload("file_0", "Reply Memo")
    elif type_of_submission == "c":
        not_dis = Select(doc_type0)
        not_dis.select_by_visible_text(
            "IMM - NOTICE OF DISCONTINUANCE / WITHDRAWAL")
        upload("file_0", "Notice of Discontinuance")
    elif type_of_submission == "d":
        not_change = Select(doc_type0)
        not_change.select_by_visible_text(
            "IMM - NOTICE OF CHANGE OF SOLICITOR / INTENTION TO ACT IN PERSON")
        upload("file_0", "Notice of Change of Solicitor")
    elif type_of_submission == "e":
        book_of_auth = Select(doc_type0)
        book_of_auth.select_by_visible_text("IMM - AUTHORITIES")
        upload("file_0", "Book of Authorities")
    elif type_of_submission == "f":
        motion_record = Select(doc_type0)
        motion_record.select_by_visible_text("IMM - MOTION RECORD")
        upload("file_0", "Motion Record")
    elif type_of_submission == "g":
        letter = Select(doc_type0)
        letter.select_by_visible_text("IMM - LETTER")
        upload("file_0", "Letter")


def upload_consent():
    # UPLOAD THE NOTICE OF CONSENT
    doc_type1 = browser.find_element_by_id("DocumentType_1")
    try:
        doc_type1.click()
        consent = Select(doc_type1)
        consent.select_by_visible_text("IMM - CONSENT")
        upload("file_1", "Notice of Consent")
    except:
        time.sleep(3)
        upload_consent()


def upload_affidavit():
    # UPLOAD THE AFF. OF SERVICE
    doc_type2 = browser.find_element_by_id("DocumentType_2")
    try:
        doc_type2.click()
        affidavit_of_service = Select(doc_type2)
        affidavit_of_service.select_by_visible_text(
            "IMM - AFFIDAVIT OF SERVICE")
        upload("file_2", "Affidavit of Service")
    except:
        time.sleep(3)
        upload_affidavit()

    # AGREE TO KEEP SIGNED COPIES
    find_click("Signatures")


def handle_glitch(type_of_submission):
    # SOMETIMES THE 'CONTINUE' BUTTON DOESN'T WORK. SO WE HAVE TO RETRY
    try:
        find_click("Submit-Step-3")
    except:
        print("The website has a glitch. I'm retrying to upload the submission")
        upload_submission(type_of_submission)
        handle_glitch(type_of_submission)


def filing_information(alt_email):
    try:
        fill_out("firstName", "TUGRUL")
        fill_out("lastName", "PINAR")
        fill_out("Address", "150 Carlton Street Suite 100")
        fill_out("City", "Toronto")
        drop_down_selection("provinceDDL", "Ontario")
        fill_out("postalCode", "M5A 2K1")
        fill_out("phoneNumber", "647-574-4320")
        fill_out("priEmail", "tugrul@lewislegal.ca")
        fill_out("secEmail", alt_email)
        drop_down_selection("languageDDL", "English")
        drop_down_selection("regOfficeDDL", "Toronto")
        find_click("Submit-Step-4")
    except:
        time.sleep(3)
        filing_information(alt_email)

    # SUBMIT
    time.sleep(1.5)
    find_click("idMyBtn")
    time.sleep(2)

    page = BeautifulSoup(browser.page_source, "html.parser")

    try:
        containers = page.findAll(
            "div", {"class": "box"})
        with open("confirmation.txt", "w") as file:
            file.write(containers[0].getText())
    except:
        with open("NOT_CONFIRMED.txt", "w") as file:
            file.write("try again")

    browser.quit()
    print("Submission confirmed")


def create_folder(doc_name, last_name):
    # CREATE A NEW FOLDER WITH APPLICANT'S LAST NAME
    os.mkdir(f"{last_name}")
    new_folder_path = Path(f"{last_name}")

    # MOVE THE PDF DOCUMENTS INSIDE NEW FOLDER
    shutil.move(f"{doc_name}.pdf", new_folder_path)
    shutil.move("Notice of Consent.pdf", new_folder_path)
    shutil.move("Affidavit of Service.pdf", new_folder_path)
    try:
        shutil.move("confirmation.txt", new_folder_path)
    except:
        pass
    print("Folder created")


def submission_date(type_of_submission, last_name):

    try:
        wb = openpyxl.load_workbook("records.xlsx")
        sheet = wb["Sheet1"]
        # PUT THE FILING DATE ON EXCEL IF WE ARE SUBMITTING AN APP. RECORD
        if type_of_submission == "a" or type_of_submission == "f":
            counter = 0
            for x in range(1, sheet.max_row + 1):
                counter += 1
                client_xl_lastname = sheet.cell(x, 1)
                cell_date = sheet.cell(counter, 6)
                # dis_date = sheet.cell(counter, 7)
                if client_xl_lastname.value == last_name and cell_date.value == None:
                    date_filed = sheet.cell(counter, 6)
                    dt = datetime.now()
                    date_filed.value = dt.strftime("%B %d, %Y")
                    print("Entry has been made")
                    break
                else:
                    continue

        # PUT THE DATE OF DISCONTINUANCE
        elif type_of_submission == "c":
            counter = 0
            for x in range(1, sheet.max_row + 1):
                counter += 1
                client_xl_lastname = sheet.cell(x, 1)
                cell_date = sheet.cell(counter, 7)
                if client_xl_lastname.value == last_name and cell_date.value == None:
                    date_filed = sheet.cell(counter, 7)
                    dt = datetime.now()
                    date_filed.value = dt.strftime("%B %d, %Y")
                    print("Entry has been made")
                    break
                else:
                    continue

        wb.save("records.xlsx")
    except:
        with open("NO_ENTRY.txt", "w") as file:
            file.write("No entry has been made")


def make_submission():
    file_no = engine.dm_imm
    type_of_submission = engine.type_letter
    app_name = engine.dm_lname.upper()
    doc_name = engine.sub_type.choices[engine.type_letter]
    existing_filing()
    court_case(file_no)
    edit_party()
    add_document()
    upload_submission(type_of_submission)
    upload_consent()
    upload_affidavit()
    handle_glitch(type_of_submission)
    filing_information(engine.dm_alt_email)
    create_folder(doc_name, app_name)
    submission_date(type_of_submission, app_name)


make_submission()
