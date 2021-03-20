from selenium import webdriver
from selenium.webdriver.support.ui import Select
import time
import openpyxl
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
import xlsxwriter as xl
from pathlib import Path
import os
import shutil
from uinput import UserInput
from decision_maker import DecisionMaker


source = os.getcwd()
dm = DecisionMaker()
dm.start_engine()

browser = webdriver.Chrome()
browser.get("https://efiling.fct-cf.gc.ca/en/online-access/e-filing-intro")
# browser.minimize_window()
browser.implicitly_wait(10)

# FIND THE WEB ELEMENT AND CLICK


def find_click(element_id):
    variable = browser.find_element_by_id(element_id)
    variable.click()

# FILL OUT THE INPUT BOXES


def fill_out(element_id, data):
    variable = browser.find_element_by_id(element_id)
    variable.send_keys(data)

# SELECT THE ITEM FROM THE DROP DOWN LIST


def drop_down_selection(element_id, selection):
    variable = browser.find_element_by_id(element_id)
    var = Select(variable)
    var.select_by_visible_text(selection)


# INITIATE A NEW APPLICATION
def new_efiling():
    find_click("disclamer")
    find_click("NewEfiling")


# STEP 1
def step_one(type_of_decision):
    # CHOOSE ENGLISH AS LANGUAGE
    drop_down_selection("proceedingLang", "English")

    # CHOOSE IMMIGRATION AS PROCEEDING TYPE
    drop_down_selection("proceedingType", "Immigration")

    # CHOOSE PROCEEDING SUBJECT BASED ON THE USER INPUT
    proceeding_subject = browser.find_element_by_id("proceedingSubject")
    if type_of_decision == "a" or type_of_decision == "b" or type_of_decision == "c" or type_of_decision == "d":
        refugee = Select(proceeding_subject)
        refugee.select_by_visible_text("Refugee")
        proceeding_nature = browser.find_element_by_id("proceedingNature")
        negative_decision = Select(proceeding_nature)
        if type_of_decision == "b":
            negative_decision.select_by_visible_text(
                "Imm - Appl. for leave & jud. review - IRB - Refugee Appeal Division")
        elif type_of_decision == "a":
            negative_decision.select_by_visible_text(
                "Imm - Appl. for leave & jud. review - IRB - Refugee Protection Div.")
        else:
            negative_decision.select_by_visible_text(
                "Imm - Appl. for leave & jud. review - Pre-removal risk assessment")
    elif type_of_decision == "e" or type_of_decision.lower() == "f":
        non_refugee = Select(proceeding_subject)
        non_refugee.select_by_visible_text("Non-Refugee")
        proceeding_nature = browser.find_element_by_id("proceedingNature")
        negative_decision = Select(proceeding_nature)
        negative_decision.select_by_visible_text(
            "Imm - Appl. for leave & jud. review - Arising outside Canada")
        decision_maker = browser.find_element_by_id("DecisionMaker")

    # CHOOSE THE DECISION MAKER BASED ON THE TYPE OF APPLICATION WE ARE APPEALING
    decision_maker = browser.find_element_by_id("DecisionMaker")
    irb = Select(decision_maker)

    if type_of_decision == type_of_decision == "d" or type_of_decision == "e" or type_of_decision == "f":
        irb.select_by_visible_text("Citizenship and Immigration Canada")

    elif type_of_decision == "a" or type_of_decision == "b":
        irb.select_by_visible_text("Immigration and Refugee Board")

    elif type_of_decision == "c":
        irb.select_by_visible_text("Canada Border Services Agency")

    find_click("Submit-Step-1")


# STEP 2
def step_two(number_of_applicants, first_names, last_names):
    # CHOOSE THE PARTY ROLE AS APPLICANT
    party_role = browser.find_element_by_id("PartyRoleMain")
    party_role.click()
    applicant = Select(party_role)
    applicant.select_by_visible_text("Applicant")

    # CHOOSE INDIVIDUAL
    drop_down_selection("PartyTypeMain", "Individual")

    # PUT FULL NAMES BASED ON THE NUMBER OF APPLICANTS
    fill_out("firstNameMain", first_names[0])
    fill_out("lastNameMain", last_names[0])

    # CLICK ADD PARTY BUTTON
    for x in range(number_of_applicants):
        find_click("addrow_Party")

    if number_of_applicants > 1:
        for x in range(number_of_applicants - 1):
            party_role_others = browser.find_element_by_id(
                f"PartyRole_{x}")
            party_role_others.click()
            applicant_others = Select(party_role_others)
            applicant_others.select_by_visible_text("Applicant")
            party_details_others = browser.find_element_by_id(
                f"PartyType_{x}")
            party_details_others.click()
            individual_others = Select(party_details_others)
            individual_others.select_by_visible_text("Individual")
            fill_out(f"firstName_{x}", first_names[x+1])
            fill_out(f"lastName_{x}", last_names[x+1])
        multiple_doj = browser.find_element_by_id(
            f"PartyRole_{number_of_applicants - 1}")
        multiple_doj.click()
        multiple_respondent = Select(multiple_doj)
        multiple_respondent.select_by_visible_text("Respondent (application)")
        doj_multiple_party_details = browser.find_element_by_id(
            f"PartyType_{number_of_applicants - 1}")
        other_multiple = Select(doj_multiple_party_details)
        other_multiple.select_by_visible_text("Other")
        fill_out(f"firstName_{number_of_applicants - 1}",
                 "The Minister of Citizenship and Immigration")
    else:
        # ADD THE MINISTER / DOJ
        drop_down_selection("PartyRole_0", "Respondent (application)")
        drop_down_selection("PartyType_0", "Other")
        fill_out("firstName_0", "The Minister of Citizenship and Immigration")

    find_click("Submit-Step-2")


# STEP 3
def step_three(number_of_applicants):
    # ADD DOCUMENT
    find_click("addrow")

    # SELECT THE DOCUMENT TYPE
    drop_down_selection(
        "DocumentType_0", "IMM - APPLICATION FOR LEAVE AND JUDICIAL REVIEW")

    # SELECT THE DOCUMENT LANGUAGE
    drop_down_selection("DocumentLanguage_0", "English")

    # CHECK ALL THE BOXES
    for x in range(number_of_applicants + 1):
        find_click(f"filer_0_{x}")

    # ATTACH THE DOCUMENT
    upload_document = browser.find_element_by_id("file_0")
    try:
        upload_document.send_keys(source + "\\App. for Leave.pdf")
        time.sleep(2)
    except Exception as e:
        print(str(e))
        print("Cannot find the document")

    find_click("Submit-Step-3")


# STEP 4
def step_four(alt_email):
    # FILING INFORMATION
    fill_out("firstName", "TUGRUL")
    fill_out("lastName", "PINAR")
    fill_out("Address", "150 Carlton Street Suite 100")
    fill_out("City", "Toronto")
    drop_down_selection("provinceDDL", "Ontario")
    fill_out("postalCode", "M5A 2K1")
    fill_out("phoneNumber", "647-574-4320")
    fill_out("priEmail", "tugrul@lewislegal.ca")
    fill_out("secEmail", alt_email)
    drop_down_selection("languageDDL", "English")
    drop_down_selection("regOfficeDDL", "Toronto")
    find_click("Submit-Step-4")

    time.sleep(1.5)

    # SUBMIT
    find_click("idMyBtn")
    time.sleep(2)

    page = BeautifulSoup(browser.page_source, "html.parser")

    try:
        cas_section = page.findAll("div", {"id": "ModalPrintArea"})
        cas_number = cas_section.findAll("div", {"class": "box"})

        with open("cas_confirmation.txt", "w") as file:
            file.write(cas_number[0].getText())
    except:
        pass

    try:
        # CONFIRM THAT THE SUBMISSION HAS BEEN MADE
        # PARSE THE HTML CONTENT TO FIND THE CONFIRMATION NUMBER
        containers = page.findAll("div", {"class": "box"})

        with open("confirmation.txt", "w") as file:
            file.write(containers[0].getText())
    except:
        with open("NOT_CONFIRMED.txt", "w") as file:
            file.write("try again")

    browser.quit()


def create_spreadsheet(first_names, last_names):
    # CREATE THE SPREADSHEET IF DOESN'T ALREADY EXISTS
    records = Path("records.xlsx")

    if not records.exists():
        wb = xl.Workbook("records.xlsx")
        bold_format = wb.add_format({"bold": True})
        cell_format = wb.add_format()
        cell_format.set_text_wrap()
        cell_format.set_align("top")
        cell_format.set_align("left=")

        sheet = wb.add_worksheet()
        sheet.write("A1", "LASTNAME", bold_format)
        sheet.write("B1", "FIRST NAME", bold_format)
        sheet.write("C1", "DATE FILED", bold_format)
        sheet.write("D1", "DUE ADMIN BY", bold_format)
        sheet.write("E1", "DUE FC BY", bold_format)
        wb.close()

    # WRITE TO THE EXCEL FILE
    wb = openpyxl.load_workbook("records.xlsx")
    sheet = wb["Sheet1"]

    for x in range(sheet.max_row+1, sheet.max_row + 2):
        client_xl_lastname = sheet.cell(x, 1)
        client_xl_lastname.value = last_names[0]
        client_xl_firstname = sheet.cell(x, 2)
        client_xl_firstname.value = first_names[0]
        date_filed = sheet.cell(x, 3)
        dt1 = datetime.now()
        date_filed.value = dt1.strftime("%A-%B %d, %Y")
        due_admin_by = sheet.cell(x, 4)
        dt2 = datetime.now() + timedelta(days=27)
        dt2 = dt2.strftime("%A-%B %d, %Y")

        # IF THE DUE DATE FALLS IN THE WEEKEND, PULL IT TO BACK FRIDAY
        if dt2.startswith("Sunday"):
            dt2 = datetime.now() + timedelta(days=25)
            due_admin_by.value = dt2.strftime("%A-%B %d, %Y")
        elif dt2.startswith("Saturday"):
            dt2 = datetime.now() + timedelta(days=25)
            due_admin_by.value = dt2.strftime("%A-%B %d, %Y")
        else:
            due_admin_by.value = dt2

        due_fc_by = sheet.cell(x, 5)
        dt3 = datetime.now() + timedelta(days=30)
        dt3 = dt3.strftime("%A-%B %d, %Y")

        # IF THE DUE DATE FALLS IN THE WEEKEND, PUSH IT TO MONDAY
        if dt3.startswith("Saturday"):
            dt3 = datetime.now() + timedelta(days=32)
            due_fc_by.value = dt3.strftime("%A-%B %d, %Y")
        elif dt3.startswith("Sunday"):
            dt3 = datetime.now() + timedelta(days=31)
            due_fc_by.value = dt3.strftime("%A-%B %d, %Y")
        else:
            due_fc_by.value = dt3

    wb.save(source + "\\records.xlsx")


def create_folder(first_names, last_names):
    # CREATE NEW FOLDER
    os.mkdir(f"{last_names[0]}, {first_names[0]}")
    new_folder_path = Path(f"{last_names[0]}, {first_names[0]}")

    # MOVE THE JR NOTICE INSIDE NEW FOLDER
    shutil.move("App. for Leave.pdf", new_folder_path)
    try:
        shutil.move("confirmation.txt", new_folder_path)
    except:
        pass


def efile_jr_notice(number_of_applicants, first_names, last_names, type_of_decision, alt_email):
    new_efiling()
    step_one(type_of_decision)
    step_two(number_of_applicants, first_names, last_names)
    step_three(number_of_applicants)
    step_four(alt_email)
    create_spreadsheet(first_names, last_names)
    create_folder(first_names, last_names)


efile_jr_notice(dm.app_number_of_applicants, dm.app_first_names,
                dm.app_last_names, dm.app_type_of_decision, dm.app_alt_email)
